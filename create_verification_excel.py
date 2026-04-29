"""
Generates KPI_Verification_Reference.xlsx in the KPIAI project root.

Three sheets:
  1. Test Data           — 15 line items, current + prior year values
  2. KPI Calculations    — step-by-step workings for all 12 KPIs
  3. Summary Reference   — compact quick-reference table
"""

import os
from datetime import date
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

OUTPUT = os.path.join(os.path.dirname(os.path.abspath(__file__)),
                      "KPI_Verification_Reference.xlsx")

# ── Palette ───────────────────────────────────────────────────────────────────
NAVY        = "1F3864"   # dark navy   — sheet title / main headers
MID_BLUE    = "2F5496"   # mid blue    — KPI section headers
STEEL       = "4472C4"   # steel blue  — formula label rows
ALT         = "DCE6F1"   # light blue  — alternating data rows
RESULT_BG   = "E2EFDA"   # light green — expected result rows
NOTE_BG     = "FFF2CC"   # pale yellow — interpretation notes
SUBTOTAL_BG = "F2F2F2"   # light grey  — sheet subtitles / column headers
WHITE       = "FFFFFF"
BLACK       = "000000"

# ── Style helpers ─────────────────────────────────────────────────────────────

def _side(color="BBBBBB", style="thin"):
    return Side(style=style, color=color)

def _border(color="BBBBBB"):
    s = _side(color)
    return Border(left=s, right=s, top=s, bottom=s)

def _thick_border():
    outer = _side("888888", "medium")
    inner = _side("BBBBBB", "thin")
    return Border(left=outer, right=outer, top=outer, bottom=outer)


def write(ws, row, col, value,
          bold=False, italic=False, size=10, color=BLACK, bg=WHITE,
          halign="left", valign="center",
          wrap=False, fmt=None, border_color="BBBBBB",
          indent=0):
    c = ws.cell(row=row, column=col, value=value)
    c.font      = Font(name="Calibri", bold=bold, italic=italic,
                       size=size, color=color)
    c.fill      = PatternFill("solid", fgColor=bg)
    c.alignment = Alignment(horizontal=halign, vertical=valign,
                            wrap_text=wrap, indent=indent)
    c.border    = _border(border_color)
    if fmt:
        c.number_format = fmt
    return c


def title_row(ws, row, ncols, text, subtitle=False):
    """Full-width title or subtitle bar."""
    bg   = SUBTOTAL_BG if subtitle else NAVY
    fg   = BLACK       if subtitle else WHITE
    sz   = 10          if subtitle else 13
    c = write(ws, row, 1, text,
              bold=not subtitle, italic=subtitle,
              size=sz, color=fg, bg=bg, halign="left",
              border_color="888888")
    if ncols > 1:
        ws.merge_cells(start_row=row, start_column=1,
                       end_row=row,   end_column=ncols)
    ws.row_dimensions[row].height = 22 if not subtitle else 16
    return c


def header_row(ws, row, cols):
    """Bold column header row on navy background."""
    for col, text in enumerate(cols, start=1):
        c = write(ws, row, col, text,
                  bold=True, size=10, color=WHITE, bg=NAVY,
                  halign="center", border_color="888888")
    ws.row_dimensions[row].height = 18


def data_row(ws, row, values, alt=False, bold=False,
             aligns=None, fmts=None, bg_override=None):
    """Standard data row with optional alternating shading."""
    bg = bg_override if bg_override else (ALT if alt else WHITE)
    aligns = aligns or ["left"] * len(values)
    fmts   = fmts   or [None]  * len(values)
    for col, (val, ha, fmt) in enumerate(zip(values, aligns, fmts), start=1):
        write(ws, row, col, val,
              bold=bold, bg=bg, halign=ha, fmt=fmt)


def set_col_widths(ws, widths):
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w


def freeze(ws, row, col=1):
    ws.freeze_panes = ws.cell(row=row, column=col)


# ── Domain data ───────────────────────────────────────────────────────────────

TEST_DATA = [
    #  Line Item                  Current   Prior
    ("Current Assets",             800.0,   750.0),
    ("Inventory",                  200.0,   180.0),
    ("Current Liabilities",        400.0,   380.0),
    ("Total Assets",             2_000.0, 1_900.0),
    ("Total Debt",                 600.0,   500.0),
    ("Total Equity",             1_000.0,   900.0),
    ("Revenue",                  1_500.0, 1_400.0),
    ("Cost of Goods Sold",         900.0,   840.0),
    ("Gross Profit",               600.0,   560.0),
    ("Operating Income",           300.0,   260.0),
    ("Operating Expenses",         300.0,   300.0),
    ("Interest Expense",            30.0,    25.0),
    ("Net Income",                 180.0,   150.0),
    ("Operating Cash Flow",        350.0,   320.0),
    ("Total Debt Service",          80.0,    70.0),
]

# Each KPI: (num, name, formula_display, [(label, content), ...], result, note or None)
KPI_DATA = [
    (
        1, "Current Ratio",
        "Current Assets  ÷  Current Liabilities",
        [
            ("Inputs",        "Current Assets = 800.0     |     Current Liabilities = 400.0"),
            ("Calculation",   "800.0  ÷  400.0  =  2.0000"),
        ],
        2.0000, None,
    ),
    (
        2, "Quick Ratio",
        "(Current Assets  −  Inventory)  ÷  Current Liabilities",
        [
            ("Inputs",              "Current Assets = 800.0     |     Inventory = 200.0     |     Current Liabilities = 400.0"),
            ("Step 1 — Quick Assets", "800.0  −  200.0  =  600.0"),
            ("Step 2 — Quick Ratio",  "600.0  ÷  400.0  =  1.5000"),
        ],
        1.5000, None,
    ),
    (
        3, "Debt-to-Equity Ratio",
        "Total Debt  ÷  Total Equity",
        [
            ("Inputs",       "Total Debt = 600.0     |     Total Equity = 1,000.0"),
            ("Calculation",  "600.0  ÷  1,000.0  =  0.6000"),
        ],
        0.6000, None,
    ),
    (
        4, "Debt-to-Asset Ratio",
        "Total Debt  ÷  Total Assets",
        [
            ("Inputs",       "Total Debt = 600.0     |     Total Assets = 2,000.0"),
            ("Calculation",  "600.0  ÷  2,000.0  =  0.3000"),
        ],
        0.3000, None,
    ),
    (
        5, "Debt-to-Equity YoY Change",
        "(Current D/E  −  Prior D/E)  ÷  Prior D/E",
        [
            ("Inputs",
             "Current: Total Debt = 600.0, Total Equity = 1,000.0     |     Prior: Total Debt = 500.0, Total Equity = 900.0"),
            ("Step 1 — Current D/E",  "600.0  ÷  1,000.0  =  0.6000"),
            ("Step 2 — Prior D/E",    "500.0  ÷  900.0  =  0.5556   (exact fraction: 5/9)"),
            ("Step 3 — YoY Change",   "(0.6000  −  0.5556)  ÷  0.5556  =  0.0444  ÷  0.5556  =  0.0800"),
            ("Exact fractions",
             "(3/5  −  5/9)  ÷  (5/9)   =   (27/45  −  25/45)  ÷  (5/9)"
             "   =   (2/45)  ×  (9/5)   =   18/225   =   2/25   =   0.08"),
        ],
        0.0800, None,
    ),
    (
        6, "Debt Service Coverage Ratio",
        "Operating Cash Flow  ÷  Total Debt Service",
        [
            ("Inputs",       "Operating Cash Flow = 350.0     |     Total Debt Service = 80.0"),
            ("Calculation",  "350.0  ÷  80.0  =  4.3750   (exact — no rounding required)"),
        ],
        4.3750, None,
    ),
    (
        7, "Return on Assets",
        "Net Income  ÷  Total Assets",
        [
            ("Inputs",       "Net Income = 180.0     |     Total Assets = 2,000.0"),
            ("Calculation",  "180.0  ÷  2,000.0  =  0.0900"),
        ],
        0.0900, None,
    ),
    (
        8, "Asset Turnover Ratio",
        "Revenue  ÷  Total Assets",
        [
            ("Inputs",       "Revenue = 1,500.0     |     Total Assets = 2,000.0"),
            ("Calculation",  "1,500.0  ÷  2,000.0  =  0.7500"),
        ],
        0.7500, None,
    ),
    (
        9, "Degree of Operating Leverage",
        "(Revenue  −  COGS  −  Operating Expenses)  ÷  Operating Income",
        [
            ("Inputs",
             "Revenue = 1,500.0     |     COGS = 900.0     |     Operating Expenses = 300.0     |     Operating Income = 300.0"),
            ("Step 1 — Numerator",  "1,500.0  −  900.0  −  300.0  =  300.0"),
            ("Step 2 — DOL",        "300.0  ÷  300.0  =  1.0000"),
        ],
        1.0000,
        "Interpretation: The numerator (Revenue − COGS − Operating Expenses = 300.0) equals Operating Income (300.0) in this "
        "dataset, so DOL = 1. This is internally consistent — it means Gross Profit minus OpEx exactly equals the reported "
        "Operating Income figure.",
    ),
    (
        10, "COGS to Revenue YoY Change",
        "(Current COGS/Revenue  −  Prior COGS/Revenue)  ÷  Prior COGS/Revenue",
        [
            ("Inputs",
             "Current: COGS = 900.0, Revenue = 1,500.0     |     Prior: COGS = 840.0, Revenue = 1,400.0"),
            ("Step 1 — Current COGS ratio",  "900.0  ÷  1,500.0  =  0.6000   (exact: 3/5)"),
            ("Step 2 — Prior COGS ratio",    "840.0  ÷  1,400.0  =  0.6000   (exact: 3/5)"),
            ("Step 3 — YoY Change",          "(0.6000  −  0.6000)  ÷  0.6000  =  0.0000"),
        ],
        0.0000,
        "Interpretation: Both COGS/Revenue ratios are exactly 60.00%. COGS scaled proportionally with Revenue "
        "in this dataset (900/1,500 = 840/1,400 = 3/5), so there is zero change in the ratio year-over-year.",
    ),
    (
        11, "Revenue Growth Rate",
        "(Current Revenue  −  Prior Revenue)  ÷  Prior Revenue",
        [
            ("Inputs",          "Current Revenue = 1,500.0     |     Prior Revenue = 1,400.0"),
            ("Calculation",     "(1,500.0  −  1,400.0)  ÷  1,400.0  =  100.0  ÷  1,400.0  =  0.071428…   (exact: 1/14)"),
            ("Rounded to 4 dp", "0.0714"),
        ],
        0.0714, None,
    ),
    (
        12, "Operating Expense Ratio YoY Change",
        "(Current OpEx/Revenue  −  Prior OpEx/Revenue)  ÷  Prior OpEx/Revenue",
        [
            ("Inputs",
             "Current: OpEx = 300.0, Revenue = 1,500.0     |     Prior: OpEx = 300.0, Revenue = 1,400.0"),
            ("Step 1 — Current OpEx ratio",  "300.0  ÷  1,500.0  =  0.2000   (exact: 1/5)"),
            ("Step 2 — Prior OpEx ratio",    "300.0  ÷  1,400.0  =  0.2143   (exact: 3/14)"),
            ("Step 3 — YoY Change",
             "(0.2000  −  0.2143)  ÷  0.2143  =  −0.0143  ÷  0.2143  =  −0.0667   (recurring)"),
            ("Exact fractions",
             "(1/5  −  3/14)  ÷  (3/14)   =   (14/70  −  15/70)  ÷  (3/14)"
             "   =   (−1/70)  ×  (14/3)   =   −14/210   =   −1/15   =   −0.0667"),
        ],
        -0.0667,
        "Interpretation: The OpEx/Revenue ratio fell 6.67% YoY because Revenue grew (dividing by 1,500 vs 1,400) "
        "while OpEx stayed flat (300.0 = 300.0). A negative result indicates improving cost efficiency relative to revenue.",
    ),
]

SUMMARY_DATA = [
    #  #    Name                                    Formula Inputs                             Result
    ( 1, "Current Ratio",                          "800 ÷ 400",                               2.0000),
    ( 2, "Quick Ratio",                            "(800 − 200) ÷ 400",                        1.5000),
    ( 3, "Debt-to-Equity Ratio",                   "600 ÷ 1,000",                              0.6000),
    ( 4, "Debt-to-Asset Ratio",                    "600 ÷ 2,000",                              0.3000),
    ( 5, "Debt-to-Equity YoY Change",              "(0.6000 − 0.5556) ÷ 0.5556",              0.0800),
    ( 6, "Debt Service Coverage Ratio",            "350 ÷ 80",                                4.3750),
    ( 7, "Return on Assets",                       "180 ÷ 2,000",                              0.0900),
    ( 8, "Asset Turnover Ratio",                   "1,500 ÷ 2,000",                            0.7500),
    ( 9, "Degree of Operating Leverage",           "(1,500 − 900 − 300) ÷ 300",               1.0000),
    (10, "COGS to Revenue YoY Change",             "(0.6000 − 0.6000) ÷ 0.6000",              0.0000),
    (11, "Revenue Growth Rate",                    "(1,500 − 1,400) ÷ 1,400",                  0.0714),
    (12, "Operating Expense Ratio YoY Change",     "(0.2000 − 0.2143) ÷ 0.2143",             -0.0667),
]


# ── Sheet builders ────────────────────────────────────────────────────────────

def build_test_data(ws):
    ws.title = "Test Data"
    ws.sheet_view.showGridLines = False

    # Sheet title
    title_row(ws, 1, 3, "KPI-AI  —  Verification Test Dataset")
    title_row(ws, 2, 3,
              f"Unit test values used for all KPI calculations  |  Generated {date.today().isoformat()}",
              subtitle=True)

    # Column headers
    header_row(ws, 3, ["Line Item", "Current Year Value", "Prior Year Value"])
    freeze(ws, 4)

    NUM_FMT = "#,##0.0"
    ALIGNS  = ["left", "right", "right"]
    FMTS    = [None, NUM_FMT, NUM_FMT]

    for i, (name, cur, prior) in enumerate(TEST_DATA, start=1):
        r   = i + 3
        alt = (i % 2 == 0)
        data_row(ws, r, [name, cur, prior], alt=alt, aligns=ALIGNS, fmts=FMTS)
        ws.row_dimensions[r].height = 16

    # Totals-style separator below last row
    last = len(TEST_DATA) + 4
    for col in range(1, 4):
        c = ws.cell(row=last, column=col, value="")
        c.fill   = PatternFill("solid", fgColor=SUBTOTAL_BG)
        c.border = _border("888888")

    set_col_widths(ws, [32, 22, 22])
    ws.row_dimensions[1].height = 24
    ws.row_dimensions[3].height = 18


def build_kpi_calculations(ws):
    ws.title = "KPI Calculations"
    ws.sheet_view.showGridLines = False

    # Sheet title
    title_row(ws, 1, 2, "KPI-AI  —  Step-by-Step KPI Calculations")
    title_row(ws, 2, 2,
              f"All 12 KPIs calculated using the verification test dataset  |  Generated {date.today().isoformat()}",
              subtitle=True)

    r = 4   # start row (row 3 is blank spacer)

    for (num, name, formula, steps, result, note) in KPI_DATA:

        # ── KPI section header ─────────────────────────────────────────────
        c = write(ws, r, 1,
                  f"KPI {num:>2}   |   {name}",
                  bold=True, size=11, color=WHITE, bg=MID_BLUE,
                  halign="left", border_color="888888")
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=2)
        ws.row_dimensions[r].height = 20
        r += 1

        # ── Formula row ────────────────────────────────────────────────────
        write(ws, r, 1, "Formula",
              bold=True, size=10, color=WHITE, bg=STEEL,
              halign="left", border_color="888888")
        write(ws, r, 2, formula,
              bold=True, size=10, color=WHITE, bg=STEEL,
              halign="left", wrap=True, border_color="888888")
        ws.row_dimensions[r].height = 18
        r += 1

        # ── Step rows ──────────────────────────────────────────────────────
        for s_idx, (label, content) in enumerate(steps):
            alt = (s_idx % 2 == 0)
            write(ws, r, 1, label,
                  bold=True, bg=ALT if alt else WHITE,
                  halign="left", border_color="BBBBBB")
            write(ws, r, 2, content,
                  bg=ALT if alt else WHITE,
                  halign="left", wrap=True, border_color="BBBBBB")
            ws.row_dimensions[r].height = 16
            r += 1

        # ── Expected result row ────────────────────────────────────────────
        write(ws, r, 1, "Expected Result (4 dp)",
              bold=True, bg=RESULT_BG, halign="left", border_color="888888")
        write(ws, r, 2, result,
              bold=True, size=12, bg=RESULT_BG,
              halign="left", fmt="0.0000", border_color="888888")
        ws.row_dimensions[r].height = 20
        r += 1

        # ── Interpretation note (if any) ───────────────────────────────────
        if note:
            c = write(ws, r, 1, note,
                      italic=True, size=9, color="5A5A5A", bg=NOTE_BG,
                      halign="left", valign="top", wrap=True,
                      border_color="BBBBBB")
            ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=2)
            ws.row_dimensions[r].height = 42
            r += 1

        # ── Blank separator ────────────────────────────────────────────────
        for col in (1, 2):
            ws.cell(row=r, column=col).value = None
        ws.row_dimensions[r].height = 8
        r += 1

    set_col_widths(ws, [34, 78])
    ws.row_dimensions[1].height = 24


def build_summary(ws):
    ws.title = "Summary Reference"
    ws.sheet_view.showGridLines = False

    # Sheet title
    title_row(ws, 1, 4, "KPI-AI  —  Summary Reference Table")
    title_row(ws, 2, 4,
              f"Quick-reference: all 12 KPIs with formula inputs and expected results  |  Generated {date.today().isoformat()}",
              subtitle=True)

    # Column headers
    header_row(ws, 3, ["#", "KPI Name", "Formula Inputs", "Expected Result (4 dp)"])
    freeze(ws, 4)

    ALIGNS = ["center", "left", "left", "center"]
    FMTS   = [None, None, None, "0.0000"]

    for i, (num, name, inputs, result) in enumerate(SUMMARY_DATA, start=1):
        r   = i + 3
        alt = (i % 2 == 0)
        data_row(ws, r, [num, name, inputs, result],
                 alt=alt, aligns=ALIGNS, fmts=FMTS)
        ws.row_dimensions[r].height = 18

    # Closing border row
    last = len(SUMMARY_DATA) + 4
    for col in range(1, 5):
        c = ws.cell(row=last, column=col, value="")
        c.fill   = PatternFill("solid", fgColor=SUBTOTAL_BG)
        c.border = _border("888888")

    set_col_widths(ws, [8, 38, 42, 24])
    ws.row_dimensions[1].height = 24
    ws.row_dimensions[3].height = 18


# ── Main ──────────────────────────────────────────────────────────────────────

def main():
    wb = Workbook()

    # openpyxl creates a default sheet; rename and reuse for Sheet 1
    ws1 = wb.active
    build_test_data(ws1)

    ws2 = wb.create_sheet()
    build_kpi_calculations(ws2)

    ws3 = wb.create_sheet()
    build_summary(ws3)

    # Set tab colours
    ws1.sheet_properties.tabColor = "1F3864"
    ws2.sheet_properties.tabColor = "2F5496"
    ws3.sheet_properties.tabColor = "4472C4"

    wb.save(OUTPUT)
    print(f"Saved → {OUTPUT}")


if __name__ == "__main__":
    main()
